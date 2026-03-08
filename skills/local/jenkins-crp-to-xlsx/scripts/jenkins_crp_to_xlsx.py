#!/usr/bin/env python3
"""
Jenkins CRP Info Collector to Excel Converter
将 Jenkins v25-crp_info_collect 任务的输出转换为 Excel 文件
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import re
import sys
import json
import subprocess
import os


def parse_jenkins_output(text):
    """解析 Jenkins 输出文本，提取主题信息"""
    themes = []
    current_theme = {}
    
    lines = text.strip().split('\n')
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        if line.startswith('主题ID:'):
            if current_theme and 'name' in current_theme:
                themes.append(current_theme)
            current_theme = {'id': line.replace('主题ID:', '').strip()}
            
        elif line.startswith('主题名称:'):
            current_theme['name'] = line.replace('主题名称:', '').strip()
            
        elif line.startswith('主题地址:'):
            current_theme['url'] = line.replace('主题地址:', '').strip()
            i += 1
            # 检查下一行是否也是 URL
            while i < len(lines) and lines[i].strip().startswith('https://crp.uniontech.com'):
                current_theme['url'] += '\n' + lines[i].strip()
                i += 1
            i -= 1
            
        elif line.startswith('仓库地址:'):
            current_theme['repo'] = line.replace('仓库地址:', '').strip()
            i += 1
            # 检查下一行是否也是仓库地址
            while i < len(lines) and lines[i].strip().startswith('deb [trusted=yes]'):
                current_theme['repo'] += '\n' + lines[i].strip()
                i += 1
            i -= 1
            
        elif line.startswith('仓库源码:'):
            current_theme['source'] = line.replace('仓库源码:', '').strip()
            i += 1
            # 收集多行源码
            while i < len(lines):
                next_line = lines[i].strip()
                if next_line.startswith('责任人:') or next_line.startswith('构建状态:') or \
                   next_line.startswith('分支:') or next_line.startswith('状态:') or \
                   next_line.startswith('主题ID:'):
                    break
                if next_line and '  ' in next_line:
                    current_theme['source'] += '\n' + next_line
                i += 1
            i -= 1
            
        elif line.startswith('责任人:'):
            current_theme['owner'] = line.replace('责任人:', '').strip()
            
        i += 1
    
    if current_theme and 'name' in current_theme:
        themes.append(current_theme)
    
    return themes


def create_excel(themes, output_path):
    """创建 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CRP-Info"
    
    # 设置表头
    headers = ['主题名称', '主题地址', '仓库', '源码名及版本', '责任人']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 填充数据
    for row_idx, theme in enumerate(themes, 2):
        ws.cell(row=row_idx, column=1, value=theme.get('name', ''))
        ws.cell(row=row_idx, column=2, value=theme.get('url', ''))
        ws.cell(row=row_idx, column=3, value=theme.get('repo', ''))
        ws.cell(row=row_idx, column=4, value=theme.get('source', ''))
        ws.cell(row=row_idx, column=5, value=theme.get('owner', ''))
        
        # 设置自动换行
        for col in range(1, 6):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # 设置列宽
    ws.column_dimensions['A'].width = 50  # 主题名称
    ws.column_dimensions['B'].width = 60  # 主题地址
    ws.column_dimensions['C'].width = 80  # 仓库
    ws.column_dimensions['D'].width = 60  # 源码名及版本
    ws.column_dimensions['E'].width = 40  # 责任人
    
    # 设置行高
    for row in range(2, len(themes) + 2):
        ws.row_dimensions[row].height = 120
    
    wb.save(output_path)
    return len(themes)


def get_jenkins_log(job_path, build_number, jk_binary='/tmp/jk'):
    """使用 jk CLI 获取 Jenkins 构建日志"""
    try:
        result = subprocess.run(
            [jk_binary, 'log', job_path, str(build_number)],
            capture_output=True,
            text=True,
            timeout=120
        )
        if result.returncode == 0:
            return result.stdout
        else:
            print(f"Error getting Jenkins log: {result.stderr}", file=sys.stderr)
            return None
    except Exception as e:
        print(f"Exception getting Jenkins log: {e}", file=sys.stderr)
        return None


def main():
    """主函数"""
    if len(sys.argv) < 2:
        print("Usage: jenkins_crp_to_xlsx.py <jenkins-job-path> [build-number] [output-path]", file=sys.stderr)
        print("Example: jenkins_crp_to_xlsx.py v25-repo-iso-work/v25-crp_info_collect 173 ./output.xlsx", file=sys.stderr)
        sys.exit(1)
    
    job_path = sys.argv[1]
    build_number = sys.argv[2] if len(sys.argv) > 2 else None
    output_path = sys.argv[3] if len(sys.argv) > 3 else None
    
    # 如果没有指定构建号，获取最新的
    if build_number is None:
        try:
            result = subprocess.run(
                ['/tmp/jk', 'run', 'ls', job_path],
                capture_output=True,
                text=True,
                timeout=30
            )
            if result.returncode == 0 and result.stdout:
                # 解析第一行获取最新的构建号
                first_line = result.stdout.strip().split('\n')[0]
                build_number = first_line.split()[0].replace('#', '')
                print(f"Using latest build: #{build_number}")
            else:
                print("Failed to get latest build number", file=sys.stderr)
                sys.exit(1)
        except Exception as e:
            print(f"Error getting latest build: {e}", file=sys.stderr)
            sys.exit(1)
    
    # 获取 Jenkins 日志
    print(f"Fetching Jenkins log for {job_path} #{build_number}...")
    log_content = get_jenkins_log(job_path, build_number)
    
    if log_content is None:
        print("Failed to fetch Jenkins log", file=sys.stderr)
        sys.exit(1)
    
    # 解析主题信息
    print("Parsing theme information...")
    themes = parse_jenkins_output(log_content)
    
    if not themes:
        print("No themes found in Jenkins output", file=sys.stderr)
        sys.exit(1)
    
    print(f"Found {len(themes)} themes")
    
    # 生成输出路径
    if output_path is None:
        # 从 job_path 提取名称
        job_name = job_path.replace('/', '-')
        output_path = f"{job_name}-{build_number}.xlsx"
    
    # 创建 Excel 文件
    print(f"Creating Excel file: {output_path}")
    count = create_excel(themes, output_path)
    
    print(f"Successfully created Excel file with {count} themes")
    print(f"Output: {os.path.abspath(output_path)}")
    
    # 输出 JSON 摘要
    summary = {
        "total_themes": count,
        "output_file": os.path.abspath(output_path),
        "themes": [
            {
                "name": t.get("name", ""),
                "owner": t.get("owner", ""),
                "source_count": len(t.get("source", "").split("\n")) if t.get("source") else 0
            }
            for t in themes
        ]
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
